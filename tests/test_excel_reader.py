from datetime import date, time

from openpyxl import Workbook

from calendar_import.excel_reader import ItineraryReadError, read_itinerary
from calendar_import.timezone_resolver import TimeZoneResolver


def resolver(tmp_path):
    return TimeZoneResolver(tmp_path / "locations.json")


def test_reads_itinerary_rows_and_notes(tmp_path):
    path = tmp_path / "trip.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Itinerary"
    sheet.append(["Date", "Location", "From", "To", "Link", "Details", "Confirmation"])
    sheet.append([date(2026, 8, 1), "Rome", time(9, 30), time(11, 0), "https://example.com", "Tour", "ABC123"])
    workbook.save(path)

    events = read_itinerary(path, resolver(tmp_path))

    assert len(events) == 1
    assert events[0].event_date == date(2026, 8, 1)
    assert events[0].location == "Rome"
    assert events[0].start_time == time(9, 30)
    assert events[0].end_time == time(11, 0)
    assert events[0].time_zone == "Europe/Rome"
    assert events[0].notes == [("Details", "Tour"), ("Confirmation", "ABC123")]


def test_blank_times_create_all_day_event(tmp_path):
    path = tmp_path / "trip.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Itinerary"
    sheet.append(["Date", "Location", "From", "To", "Link"])
    sheet.append(["2026-08-01", "New York", "", "", ""])
    workbook.save(path)

    events = read_itinerary(path, resolver(tmp_path))

    assert events[0].is_all_day


def test_one_blank_time_uses_the_non_blank_time_for_both_start_and_end(tmp_path):
    path = tmp_path / "trip.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Itinerary"
    sheet.append(["Date", "Location", "From", "To", "Link"])
    sheet.append(["2026-08-01", "New York", "8:15 AM", "", ""])
    workbook.save(path)

    events = read_itinerary(path, resolver(tmp_path))

    assert events[0].start_time == time(8, 15)
    assert events[0].end_time == time(8, 15)


def test_requires_itinerary_sheet(tmp_path):
    path = tmp_path / "trip.xlsx"
    Workbook().save(path)

    try:
        read_itinerary(path, resolver(tmp_path))
    except ItineraryReadError as exc:
        assert "Itinerary" in str(exc)
    else:
        raise AssertionError("Expected an ItineraryReadError")

