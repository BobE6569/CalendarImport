from __future__ import annotations

from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import ItineraryEvent
from .timezone_resolver import TimeZoneResolver


REQUIRED_HEADERS = ["Date", "Location", "From", "To", "Link"]


class ItineraryReadError(ValueError):
    pass


def read_itinerary(path: Path, time_zone_resolver: TimeZoneResolver) -> list[ItineraryEvent]:
    try:
        workbook = load_workbook(path, data_only=True)
    except PermissionError as exc:
        raise ItineraryReadError(
            "CalendarImport could not open the Excel file.\n\n"
            "Close the workbook in Excel, make sure OneDrive has downloaded it locally, "
            "then try again. If it still fails, save a copy to a local folder such as "
            "Documents and select that copy."
        ) from exc
    except FileNotFoundError as exc:
        raise ItineraryReadError("The selected Excel file could not be found. Select the file again.") from exc
    except (BadZipFile, InvalidFileException, OSError) as exc:
        raise ItineraryReadError(
            "CalendarImport could not read the selected workbook. "
            "Make sure it is a valid .xlsx or .xlsm file and that it is not blocked by OneDrive or Excel."
        ) from exc
    if "Itinerary" not in workbook.sheetnames:
        raise ItineraryReadError("The workbook does not contain a tab named 'Itinerary'.")

    sheet = workbook["Itinerary"]
    headers = [_cell_text(sheet.cell(row=1, column=column).value) for column in range(1, sheet.max_column + 1)]
    if headers[:5] != REQUIRED_HEADERS:
        found = ", ".join(headers[:5])
        expected = ", ".join(REQUIRED_HEADERS)
        raise ItineraryReadError(f"The first five headers must be: {expected}. Found: {found}.")

    events: list[ItineraryEvent] = []
    previous_time_zone: str | None = None

    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_number, column=column).value for column in range(1, sheet.max_column + 1)]
        if all(_is_blank(value) for value in values):
            continue

        event_date = _parse_date(values[0], row_number)
        location = _cell_text(values[1])
        from_time = _parse_time(values[2], row_number, "From")
        to_time = _parse_time(values[3], row_number, "To")
        start_time, end_time = _normalize_times(from_time, to_time)
        url = _cell_text(values[4])
        time_zone = time_zone_resolver.resolve(location, previous_time_zone)
        previous_time_zone = time_zone

        notes = []
        for header, value in zip(headers[5:], values[5:]):
            label = _cell_text(header)
            content = _cell_text(value)
            if label and content:
                notes.append((label, content))

        events.append(
            ItineraryEvent(
                row_number=row_number,
                event_date=event_date,
                location=location,
                start_time=start_time,
                end_time=end_time,
                url=url,
                notes=notes,
                time_zone=time_zone,
            )
        )

    return events


def _normalize_times(from_time: time | None, to_time: time | None) -> tuple[time | None, time | None]:
    if from_time is None and to_time is None:
        return None, None
    if from_time is None:
        return to_time, to_time
    if to_time is None:
        return from_time, from_time
    return from_time, to_time


def _parse_date(value: Any, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d, %Y"):
            try:
                return datetime.strptime(cleaned, pattern).date()
            except ValueError:
                pass
    raise ItineraryReadError(f"Row {row_number}: Date is required and must be a valid date.")


def _parse_time(value: Any, row_number: int, column_name: str) -> time | None:
    if _is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.time().replace(second=0, microsecond=0)
    if isinstance(value, time):
        return value.replace(second=0, microsecond=0)
    if isinstance(value, (int, float)):
        if not 0 <= float(value) < 1:
            raise ItineraryReadError(f"Row {row_number}: {column_name} must be a time of day.")
        seconds = round(float(value) * 24 * 60 * 60)
        return (datetime.min + timedelta(seconds=seconds)).time().replace(second=0, microsecond=0)
    if isinstance(value, str):
        cleaned = value.strip()
        for pattern in ("%H:%M", "%I:%M %p", "%I %p", "%H%M"):
            try:
                return datetime.strptime(cleaned.upper(), pattern).time().replace(second=0, microsecond=0)
            except ValueError:
                pass
    raise ItineraryReadError(f"Row {row_number}: {column_name} must be blank or a valid time.")


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="minutes")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())
