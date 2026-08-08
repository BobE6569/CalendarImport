from __future__ import annotations

from datetime import time
from pathlib import Path

from openpyxl import load_workbook

from .excel_reader import ItineraryReadError, _cell_text, _is_blank, _parse_date, _parse_time
from .models import ItineraryEvent


REQUIRED_HEADERS = ["Date", "Event", "Location", "From", "To", "Reservations Open"]
PCS_EVENT_LOCATION = "8060 Grand Lely Dr Naples FL 34113 United States"
PCS_TIME_ZONE = "America/New_York"


def read_pcs_club_events(path: Path) -> list[ItineraryEvent]:
    try:
        workbook = load_workbook(path, data_only=True)
    except Exception as exc:
        raise ItineraryReadError(f"CalendarImport could not read the selected PCS workbook: {exc}") from exc

    sheet = workbook[workbook.sheetnames[0]]
    headers = [_cell_text(sheet.cell(row=1, column=column).value) for column in range(1, len(REQUIRED_HEADERS) + 1)]
    if headers != REQUIRED_HEADERS:
        expected = ", ".join(REQUIRED_HEADERS)
        found = ", ".join(headers)
        raise ItineraryReadError(f"The first PCS sheet must start with these headers: {expected}. Found: {found}.")

    events: list[ItineraryEvent] = []
    for row_number in range(2, sheet.max_row + 1):
        values = [sheet.cell(row=row_number, column=column).value for column in range(1, len(REQUIRED_HEADERS) + 1)]
        if all(_is_blank(value) for value in values):
            continue

        event_date = _parse_date(values[0], row_number)
        event_title = _cell_text(values[1])
        club_location = _cell_text(values[2])
        from_time = _parse_time(values[3], row_number, "From")
        if from_time is None:
            raise ItineraryReadError(f"Row {row_number}: From is required for PCS Club Events.")
        to_time = _parse_time(values[4], row_number, "To") or time(22, 0)
        reservations_open_date = _parse_date(values[5], row_number)
        note = _pcs_note(event_date, from_time, to_time, event_title, club_location)

        events.append(
            ItineraryEvent(
                row_number=row_number,
                event_date=reservations_open_date,
                location=club_location,
                start_time=time(9, 0),
                end_time=time(9, 0),
                url="",
                notes=[],
                time_zone=PCS_TIME_ZONE,
                title_override=f"PCS Reservation: {event_title}",
                note_override=note,
                reminder_minutes_before_start=0,
            )
        )
        events.append(
            ItineraryEvent(
                row_number=row_number,
                event_date=event_date,
                location=PCS_EVENT_LOCATION,
                start_time=from_time,
                end_time=to_time,
                url="",
                notes=[],
                time_zone=PCS_TIME_ZONE,
                title_override=f"PCS: {event_title}",
                note_override=note,
                reminder_minutes_before_start=60,
            )
        )

    return events


def _pcs_note(event_date, from_time, to_time, event_title: str, club_location: str) -> str:
    first_line = f"{event_date.isoformat()} {_format_time(from_time)} {_format_time(to_time)} {event_title}"
    return f"{first_line}\n{club_location}"


def _format_time(value: time) -> str:
    return value.strftime("%I:%M %p").lstrip("0")

